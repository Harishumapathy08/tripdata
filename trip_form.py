import streamlit as st
import pandas as pd
import os
import io
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook

st.set_page_config(layout="wide")

DATA_FILE = "data/trip_data.xlsx"
drivers = ["Prem", "Ajith", "Wilson"]
columns = [
    "S.No.", "Driver", "Disp Date", "Invoice No", "Customer", "Destination",
    "Invoice Date", "Vehicle", "Out Time", "In Time", "Out KM", "In KM", "Diff in KM"
]

def create_empty_excel_file():
    os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
    wb = Workbook()
    # Remove the default sheet created by Workbook
    default_sheet = wb.active
    wb.remove(default_sheet)
    for drv in drivers:
        ws = wb.create_sheet(title=drv)
        ws.append(columns)
    wb.save(DATA_FILE)

def load_data():
    try:
        if not os.path.exists(DATA_FILE):
            create_empty_excel_file()
        xls = pd.ExcelFile(DATA_FILE, engine='openpyxl')
    except Exception as e:
        st.warning(f"Excel file issue detected: {e}. Creating a new file.")
        create_empty_excel_file()
        xls = pd.ExcelFile(DATA_FILE, engine='openpyxl')

    all_data = pd.DataFrame(columns=columns)
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet)
        all_data = pd.concat([all_data, df], ignore_index=True)
    return all_data

def save_to_driver_sheet(driver, df):
    if os.path.exists(DATA_FILE):
        book = load_workbook(DATA_FILE)
        if driver in book.sheetnames:
            del book[driver]
        book.save(DATA_FILE)

    with pd.ExcelWriter(DATA_FILE, engine='openpyxl', mode='a' if os.path.exists(DATA_FILE) else 'w') as writer:
        df.to_excel(writer, sheet_name=driver, index=False)

def validate_time(t):
    # Validate simple hh:mm format (24-hour)
    pattern = r'^([01]\d|2[0-3]):[0-5]\d$'
    return re.match(pattern, t.strip()) is not None

st.title("🚛 Trip Entry Form")

df = load_data()

st.subheader("➕ Add Trip Record")
with st.form("trip_form"):
    selected_driver = st.selectbox("Driver", drivers)
    col1, col2, col3 = st.columns(3)
    disp_date = col1.date_input("Disp Date")
    inv_no = col2.text_input("Invoice No")
    customer = col3.text_input("Customer")

    destination = col1.text_input("Destination")
    inv_date = col2.date_input("Invoice Date")
    vehicle = col3.text_input("Vehicle")

    out_time = col1.text_input("Out Time (hh:mm, 24-hour format, e.g., 14:30)")
    in_time = col2.text_input("In Time (hh:mm, 24-hour format, e.g., 05:45)")

    out_km = col3.number_input("Out KM", 0)
    in_km = col1.number_input("In KM", 0)
    diff_km = in_km - out_km if in_km >= out_km else 0

    submitted = st.form_submit_button("Submit")

    if submitted:
        if not validate_time(out_time):
            st.error("Out Time format is invalid. Use hh:mm 24-hour format (00:00 to 23:59).")
        elif not validate_time(in_time):
            st.error("In Time format is invalid. Use hh:mm 24-hour format (00:00 to 23:59).")
        else:
            driver_df = df[df["Driver"] == selected_driver].copy()
            new_row = [
                len(driver_df) + 1,
                selected_driver,
                disp_date,
                inv_no,
                customer,
                destination,
                inv_date,
                vehicle,
                out_time.strip(),
                in_time.strip(),
                out_km,
                in_km,
                diff_km
            ]
            new_entry = pd.DataFrame([new_row], columns=columns)
            driver_df = pd.concat([driver_df, new_entry], ignore_index=True)
            save_to_driver_sheet(selected_driver, driver_df)
            st.success(f"✅ Trip added for {selected_driver}")
            df = load_data()

st.subheader("📋 View Trips")
driver_filter = st.selectbox("Select Driver to View", drivers, key="view_driver")
filtered_df = df[df["Driver"] == driver_filter]

if not filtered_df.empty:
    st.dataframe(filtered_df, use_container_width=True)

    delete_index = st.number_input("Enter S.No. to Delete", min_value=1, max_value=len(filtered_df), step=1)
    if st.button("🗑 Delete Trip"):
        filtered_df = filtered_df[filtered_df["S.No."] != delete_index]
        filtered_df.reset_index(drop=True, inplace=True)
        filtered_df["S.No."] = range(1, len(filtered_df) + 1)
        save_to_driver_sheet(driver_filter, filtered_df)
        st.success(f"🗑 Trip deleted from {driver_filter}'s sheet.")
        df = load_data()

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        for drv in drivers:
            temp_df = df[df["Driver"] == drv]
            if not temp_df.empty:
                temp_df.to_excel(writer, sheet_name=drv, index=False)
    buffer.seek(0)
    st.download_button("⬇️ Download All Trip Data", data=buffer, file_name="trip_data.xlsx")
else:
    st.info("No records found for this driver.")
